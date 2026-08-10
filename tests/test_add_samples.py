"""Tests for neotoma2faire.write.samples.add_samples.

Chronology moved out of ``sampleMetadata`` and into the ``ageModels`` sheet, so
these tests pin down that the sheet stays free of age columns.
"""

import openpyxl
import pandas as pd
import pytest

from neotoma2faire.write.samples import add_samples

HEADER_ROW = 3


def _make_wb(headers: list[str]) -> openpyxl.Workbook:
    """Return a workbook with a ``sampleMetadata`` sheet headed by *headers*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sampleMetadata"
    for col_idx, name in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=name)
    return wb


@pytest.fixture
def long_df():
    """Long-format get_data() frame: two samples × two taxa, with ages."""
    return pd.DataFrame({
        "sampleid": [1, 1, 2, 2],
        "samp_name": ["S1", "S1", "S2", "S2"],
        "taxonid": [10, 11, 10, 11],
        "value": [5, 6, 7, 8],
        "minimumDepthInMeters": [0.5, 0.5, 1.5, 1.5],
        "age": [131.7, 131.7, 500.0, 500.0],
        "ageUnit": ["Calendar years BP"] * 4,
        "agemodel": ["CRS"] * 4,
    })


class TestAddSamples:
    def test_one_row_per_sample(self, long_df):
        wb = _make_wb(["samp_name", "minimumDepthInMeters"])
        add_samples(wb, long_df, header_row=HEADER_ROW)
        ws = wb["sampleMetadata"]
        assert [ws.cell(row=r, column=1).value for r in (4, 5)] == ["S1", "S2"]
        assert ws.cell(row=6, column=1).value is None

    def test_age_columns_not_appended(self, long_df):
        """Ages belong to ageModels; sampleMetadata must not grow age columns."""
        wb = _make_wb(["samp_name", "minimumDepthInMeters"])
        add_samples(wb, long_df, header_row=HEADER_ROW)
        ws = wb["sampleMetadata"]
        headers = [c.value for c in ws[HEADER_ROW] if c.value is not None]
        assert headers == ["samp_name", "minimumDepthInMeters"]

    def test_returns_otu_pivot(self, long_df):
        wb = _make_wb(["samp_name"])
        pivot = add_samples(wb, long_df, header_row=HEADER_ROW)
        assert list(pivot.columns) == ["taxonid", "sample_1", "sample_2"]
        assert len(pivot) == 2
