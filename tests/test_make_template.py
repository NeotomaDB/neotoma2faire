"""End-to-end tests for neotoma2faire.make_template.

``make_template`` is the orchestrator: it loads the shipped FAIRe checklist,
fetches data, and hands each sheet to its writer. These tests run the real
workbook through the real writers, stubbing only the network boundary
(``get_data``, ``get_assays_by_dataset``, ``get_taxa_batch``, ``add_project``).
"""

from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from neotoma2faire.make_template import make_template

TEMPLATE = "assets/FAIRe_checklist_v1.0.2.xlsx"


def _data(extra_rows=()):
    """Two samples × two taxa, in the shape get_data returns."""
    rows = []
    for sampleid, name, depth in ((1, "S1", 0.5), (2, "S2", 1.5)):
        for taxonid, value in ((101, 12), (102, 3)):
            rows.append(
                {
                    "siteid": 5001,
                    "sitename": "Lake Okoboji",
                    "decimalLatitude": 43.38,
                    "decimalLongitude": -95.15,
                    "elev": 430,
                    "geo_loc_name": "United States, Iowa",
                    "collectionunitid": 6001,
                    "eventDate": "2018-07-14",
                    "verbatimEventDate": "2018-07-14",
                    "samp_collect_device": "Livingstone piston corer",
                    "samp_collect_method": "core WLO18",
                    "env_medium": "Lacustrine",
                    "agemodel": "CRS",
                    "modelagetype": "Calendar years BP",
                    "datasetid": 74655,
                    "datasettype": "aeDNA",
                    "sampleid": sampleid,
                    "samp_name": name,
                    "analysisunitid": 1000 + sampleid,
                    "sample_derived_from": 1000 + sampleid,
                    "analysisunitname": f"AU-{sampleid}",
                    "depth": depth,
                    "thickness": 1.0,
                    "minimumDepthInMeters": depth,
                    "maximumDepthInMeters": depth,
                    "materialSampleID": f"IGSN{sampleid}",
                    "samp_mat_process": "DNA extraction",
                    "samp_category": "sample",
                    "age": 150 * sampleid,
                    "ageOldest": 175 * sampleid,
                    "ageYoungest": 125 * sampleid,
                    "ageUnit": "Calendar years BP",
                    "taxonid": taxonid,
                    "value": value,
                    "variablename": f"taxon-{taxonid}",
                    "units": "present/absent",
                    "element": "DNA",
                    "taxongroup": "Vascular plants",
                    "ecologicalgroup": "TRSH",
                }
            )
    rows.extend(extra_rows)
    return pd.DataFrame(rows)


def _lab_row(base_row):
    """A Laboratory-analyses datum, which is filed as a taxon but is not one."""
    row = dict(base_row)
    row.update(
        {
            "taxonid": 999,
            "value": 0.14,
            "variablename": "Sedimentation rate",
            "taxongroup": "Laboratory analyses",
        }
    )
    return row


TAXA = [
    {"taxonid": 101, "taxonname": "Picea"},
    {"taxonid": 102, "taxonname": "Abies"},
]


_DEFAULT = object()  # lets a test pass output=None to mean "let make_template choose"


def _run(tmp_path, data=None, output=_DEFAULT, assays=None, taxa=None, sequences=None,
         template=TEMPLATE):
    """Invoke make_template with the network boundary stubbed out."""
    args = SimpleNamespace(
        template=template,
        dataset=74655,
        output=str(tmp_path / "out.xlsx") if output is _DEFAULT else output,
    )
    with (
        patch("neotoma2faire.make_template.get_data",
              return_value=data if data is not None else _data()),
        patch("neotoma2faire.make_template.get_assays_by_dataset",
              return_value=assays if assays is not None else []),
        patch("neotoma2faire.make_template.add_project"),
        patch("neotoma2faire.write.taxa.get_taxa_batch",
              return_value=taxa if taxa is not None else TAXA) as taxa_mock,
        # Two writers reach /aedna/sequences/{id}: add_taxa (via the dataset_id
        # make_template passes through) and add_final_reads. Stub both, or the
        # suite silently makes live HTTP calls.
        patch("neotoma2faire.write.taxa.get_taxa_sequences",
              return_value=sequences if sequences is not None else pd.DataFrame()),
        patch("neotoma2faire.write.final_reads.get_taxa_sequences",
              return_value=sequences if sequences is not None else pd.DataFrame()),
    ):
        written = make_template(args)
    return written, taxa_mock


class TestMakeTemplate:
    def test_writes_the_requested_output_path(self, tmp_path):
        out = tmp_path / "FAIRe.xlsx"
        written, _ = _run(tmp_path, output=str(out))

        assert written == str(out)
        assert out.exists()

    def test_source_template_is_not_modified(self, tmp_path):
        before = Path(TEMPLATE).read_bytes()
        _run(tmp_path)

        assert Path(TEMPLATE).read_bytes() == before

    def test_default_output_path_is_derived_from_dataset_id(self, tmp_path, monkeypatch):
        """output=None must not overwrite the source template."""
        template = Path(TEMPLATE).resolve()  # resolve before leaving the repo root
        monkeypatch.chdir(tmp_path)
        written, _ = _run(tmp_path, output=None, template=str(template))

        assert written == "outputs/FAIRe_DS_74655.xlsx"
        assert (tmp_path / "outputs" / "FAIRe_DS_74655.xlsx").exists()

    def test_all_checklist_sheets_survive(self, tmp_path):
        written, _ = _run(tmp_path)
        wb = load_workbook(written)

        for sheet in (
            "README",
            "projectMetadata",
            "sampleMetadata",
            "experimentRunMetadata",
            "taxaRaw",
            "taxaFinal",
        ):
            assert sheet in wb.sheetnames

    def test_extra_sheets_are_created(self, tmp_path):
        """ageModels and finalReads are not in the FAIRe checklist."""
        written, _ = _run(tmp_path)
        wb = load_workbook(written)

        assert "ageModels" in wb.sheetnames

    def test_sample_metadata_gets_one_row_per_sample(self, tmp_path):
        written, _ = _run(tmp_path)
        ws = load_workbook(written)["sampleMetadata"]

        header = {c.value: c.column for c in ws[3]}
        names = [
            ws.cell(row=r, column=header["samp_name"]).value
            for r in range(4, 4 + 2)
        ]
        assert names == ["S1", "S2"]

    def test_experiment_run_gets_one_row_per_sample(self, tmp_path):
        written, _ = _run(tmp_path)
        ws = load_workbook(written)["experimentRunMetadata"]

        header = {c.value: c.column for c in ws[3]}
        assert ws.cell(row=4, column=header["samp_name"]).value == "S1"
        assert ws.cell(row=5, column=header["samp_name"]).value == "S2"

    def test_taxa_sheets_receive_the_scientific_names(self, tmp_path):
        written, _ = _run(tmp_path)
        wb = load_workbook(written)

        for sheet in ("taxaFinal", "taxaRaw"):
            ws = wb[sheet]
            header = {c.value: c.column for c in ws[3]}
            names = [ws.cell(row=r, column=header["scientificName"]).value for r in (4, 5)]
            assert set(names) == {"Picea", "Abies"}

    def test_lab_analyses_are_excluded_from_the_taxa_lookup(self, tmp_path):
        """Sedimentation rate is filed as a taxon in ndb.data but is not one."""
        data = _data()
        data = pd.concat([data, pd.DataFrame([_lab_row(data.iloc[0])])], ignore_index=True)
        _, taxa_mock = _run(tmp_path, data=data)

        (requested_ids,) = taxa_mock.call_args.args
        assert sorted(requested_ids) == [101, 102]
        assert 999 not in requested_ids

    def test_taxon_ids_are_deduplicated(self, tmp_path):
        _, taxa_mock = _run(tmp_path)

        (requested_ids,) = taxa_mock.call_args.args
        assert sorted(requested_ids) == [101, 102]

    def test_creates_missing_output_directory(self, tmp_path):
        out = tmp_path / "nested" / "deeper" / "out.xlsx"
        written, _ = _run(tmp_path, output=str(out))

        assert out.exists()
        assert written == str(out)
