"""Tests for neotoma2faire.write.taxa.

``add_taxa`` is deliberately flat: it writes the leaf taxon names into both
taxaFinal and taxaRaw and leaves the Linnaean rank columns blank, because the
hierarchy walk in ``extract.taxa`` costs one REST call per ancestor. These
tests pin that contract, including the aliases the OTU merge depends on.
"""

from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook

from neotoma2faire.write.taxa import add_taxa

TAXA = [
    {"taxonid": 101, "taxonname": "Picea"},
    {"taxonid": 102, "taxonname": "Abies"},
]


def _workbook(header_row=3):
    """A minimal stand-in carrying only the columns add_taxa writes."""
    wb = Workbook()
    wb.remove(wb.active)
    columns = [
        "scientificName",
        "taxonID",
        "taxonID_db",
        "verbatimIdentification",
        "kingdom",
        "phylum",
        "seq_id",
        "dna_sequence",
    ]
    for name in ("taxaFinal", "taxaRaw"):
        ws = wb.create_sheet(name)
        for idx, column in enumerate(columns, start=1):
            ws.cell(row=header_row, column=idx, value=column)
    return wb


def _run(txid, taxa=None, wb=None, **kwargs):
    wb = wb if wb is not None else _workbook()
    with patch(
        "neotoma2faire.write.taxa.get_taxa_batch",
        return_value=taxa if taxa is not None else TAXA,
    ) as mock:
        df = add_taxa(wb, txid, **kwargs)
    return wb, df, mock


def _column(ws, name, rows, header_row=3):
    header = {c.value: c.column for c in ws[header_row]}
    return [ws.cell(row=r, column=header[name]).value for r in rows]


class TestAddTaxa:
    def test_writes_both_taxa_sheets_identically(self):
        wb, _, _ = _run([101, 102])

        for sheet in ("taxaFinal", "taxaRaw"):
            assert set(_column(wb[sheet], "scientificName", (4, 5))) == {"Picea", "Abies"}

    def test_scalar_taxon_id_is_accepted(self):
        _, _, mock = _run(101, taxa=[TAXA[0]])

        assert mock.call_args.args == ([101],)

    def test_ids_are_deduplicated_and_coerced_to_int(self):
        _, _, mock = _run([101, 101, 102.0, "102"])

        (requested,) = mock.call_args.args
        assert sorted(requested) == [101, 102]
        assert all(isinstance(i, int) for i in requested)

    def test_taxon_id_db_is_stamped_neotoma(self):
        _, df, _ = _run([101, 102])

        assert list(df["taxonID_db"]) == ["Neotoma", "Neotoma"]

    def test_verbatim_identification_mirrors_scientific_name(self):
        _, df, _ = _run([101, 102])

        assert list(df["verbatimIdentification"]) == list(df["scientificName"])

    def test_otu_merge_aliases_are_present(self):
        _, df, _ = _run([101, 102])

        assert list(df["most_specific_id"]) == list(df["taxonID"])
        assert list(df["most_specific_name"]) == list(df["scientificName"])

    def test_rank_and_sequence_columns_are_left_blank(self):
        """Neotoma's REST API exposes neither, so they must stay empty."""
        wb, _, _ = _run([101, 102])
        ws = wb["taxaFinal"]

        for column in ("kingdom", "phylum", "seq_id", "dna_sequence"):
            assert _column(ws, column, (4, 5)) == [None, None]

    def test_a_dataset_with_no_sequences_matches_the_no_dataset_id_output(self):
        """``dataset_id`` only ever adds sequences; it changes nothing else."""
        _, df_without, _ = _run([101, 102])
        with patch("neotoma2faire.write.taxa.get_taxa_sequences", return_value=pd.DataFrame()):
            _, df_with, _ = _run([101, 102], dataset_id=74655)

        assert df_with.equals(df_without)

    def test_custom_header_row(self):
        wb = _workbook(header_row=1)
        wb, _, _ = _run([101, 102], wb=wb, header_row=1)

        assert set(_column(wb["taxaFinal"], "scientificName", (2, 3), header_row=1)) == {
            "Picea",
            "Abies",
        }

    def test_empty_taxa_response_writes_nothing(self):
        wb, df, _ = _run([101], taxa=[])

        assert df.empty
        assert _column(wb["taxaFinal"], "scientificName", (4,)) == [None]


def _run_with_sequences(txid, sequences, taxa=None, **kwargs):
    """Run add_taxa with both network boundaries stubbed."""
    wb = _workbook()
    with (
        patch(
            "neotoma2faire.write.taxa.get_taxa_batch",
            return_value=taxa if taxa is not None else TAXA,
        ),
        patch("neotoma2faire.write.taxa.get_taxa_sequences", return_value=sequences) as seq_mock,
    ):
        df = add_taxa(wb, txid, **kwargs)
    return wb, df, seq_mock


def _sequences(rows):
    """A get_taxa_sequences-shaped frame carrying only the merged columns."""
    return pd.DataFrame(rows, columns=["taxonid", "sequence", "sequenceid"])


class TestAddTaxaSequences:
    """``dataset_id`` pulls the dataset's DNA sequences into both taxa sheets."""

    def test_dataset_id_fills_dna_sequence_and_seq_id(self):
        sequences = _sequences([
            {"taxonid": 101, "sequence": "ACGT", "sequenceid": 7},
            {"taxonid": 102, "sequence": "TGCA", "sequenceid": 8},
        ])
        wb, df, seq_mock = _run_with_sequences([101, 102], sequences, dataset_id=74666)

        seq_mock.assert_called_once_with(74666)
        assert sorted(df["dna_sequence"]) == ["ACGT", "TGCA"]
        for sheet in ("taxaFinal", "taxaRaw"):
            ws = wb[sheet]
            assert sorted(_column(ws, "dna_sequence", [4, 5])) == ["ACGT", "TGCA"]
        assert sorted(_column(wb["taxaFinal"], "seq_id", [4, 5])) == [7, 8]

    def test_a_taxon_with_several_sequences_yields_one_row_each(self):
        # The sequence, not the name, identifies the row — two ASVs of the same
        # taxon must both survive.
        sequences = _sequences([
            {"taxonid": 101, "sequence": "AAAA", "sequenceid": 1},
            {"taxonid": 101, "sequence": "CCCC", "sequenceid": 2},
            {"taxonid": 102, "sequence": "GGGG", "sequenceid": 3},
        ])
        _, df, _ = _run_with_sequences([101, 102], sequences, dataset_id=74666)

        assert len(df) == 3
        picea = df[df["scientificName"] == "Picea"]
        assert sorted(picea["dna_sequence"]) == ["AAAA", "CCCC"]

    def test_string_taxon_ids_from_the_api_still_merge(self):
        # The API returns IDs as strings in places; an unaligned dtype would
        # merge to all-blank instead of raising.
        sequences = _sequences([{"taxonid": "101", "sequence": "ACGT", "sequenceid": "7"}])
        _, df, _ = _run_with_sequences([101], sequences, taxa=[TAXA[0]], dataset_id=74666)

        assert df["dna_sequence"].tolist() == ["ACGT"]

    def test_no_sequences_on_record_leaves_the_columns_blank(self):
        _, df, _ = _run_with_sequences([101, 102], pd.DataFrame(), dataset_id=74666)

        assert "dna_sequence" not in df.columns
        assert len(df) == 2

    def test_without_dataset_id_sequences_are_never_fetched(self):
        wb = _workbook()
        with (
            patch("neotoma2faire.write.taxa.get_taxa_batch", return_value=TAXA),
            patch("neotoma2faire.write.taxa.get_taxa_sequences") as seq_mock,
        ):
            add_taxa(wb, [101, 102])
        seq_mock.assert_not_called()
        assert _column(wb["taxaFinal"], "dna_sequence", [4, 5]) == [None, None]
