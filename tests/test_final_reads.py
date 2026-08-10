"""Tests for neotoma2faire.extract.final_reads and write.final_reads."""

import openpyxl
import pandas as pd
import pytest

from neotoma2faire.extract.final_reads import get_final_reads
from neotoma2faire.write.final_reads import add_final_reads


@pytest.fixture
def long_df():
    """Two samples × (two sequenced taxa + one lab-analysis pseudo-taxon)."""
    rows = []
    for sampleid, samp_name, reads in [(1, "WLO17", (3860, 21)), (2, "WLO18", (6367, 122))]:
        rows.append({"sampleid": sampleid, "samp_name": samp_name, "taxonid": 69612,
                     "value": reads[0], "variablename": "Hydrurales", "units": "reads"})
        rows.append({"sampleid": sampleid, "samp_name": samp_name, "taxonid": 69613,
                     "value": reads[1], "variablename": "Choricystis", "units": "reads"})
        rows.append({"sampleid": sampleid, "samp_name": samp_name, "taxonid": 49384,
                     "value": 0.26, "variablename": "Sedimentation rate", "units": "cm/yr"})
    return pd.DataFrame(rows)


@pytest.fixture
def sequences():
    return pd.DataFrame({
        "taxonid":    [69612, 69613],
        "taxonname":  ["Hydrurales_Clade-II_X_sp.", "Choricystis_limnetica"],
        "sequenceid": [12, 13],
        "sequence":   ["GGGGAAACT", "GGGAAAACT"],
        "asv":        ["ASV1", "ASV2"],
    })


class TestGetFinalReads:
    def test_one_row_per_sequence(self, long_df, sequences):
        result = get_final_reads(long_df, sequences)
        assert len(result) == 2
        assert list(result["ASV"]) == ["ASV1", "ASV2"]

    def test_sample_name_columns_hold_counts(self, long_df, sequences):
        result = get_final_reads(long_df, sequences)
        assert list(result.columns[-2:]) == ["WLO17", "WLO18"]
        assert result.loc[0, "WLO17"] == 3860
        assert result.loc[1, "WLO18"] == 122

    def test_identity_columns(self, long_df, sequences):
        result = get_final_reads(long_df, sequences)
        assert result.loc[0, "scientificName"] == "Hydrurales_Clade-II_X_sp."
        assert result.loc[0, "DNAsequence"] == "GGGGAAACT"
        assert result.loc[0, "taxonID_db"] == "Neotoma"
        assert result.loc[0, "Units"] == "reads"
        assert result.loc[0, "verbatimIdentification"] is None

    def test_unsequenced_taxa_excluded(self, long_df, sequences):
        """The 'Sedimentation rate' pseudo-taxon belongs to ageModels, not here."""
        result = get_final_reads(long_df, sequences)
        assert 49384 not in set(sequences["taxonid"])
        assert len(result) == 2

    def test_repeated_taxon_paired_with_its_sequences(self, long_df):
        """Two sequences for one taxon get one row each, not a cross-product."""
        repeated = pd.concat([long_df, long_df[long_df["taxonid"] == 69612]])
        seqs = pd.DataFrame({
            "taxonid":    [69612, 69612],
            "taxonname":  ["Same name", "Same name"],
            "sequenceid": [12, 99],
            "sequence":   ["GGGGAAACT", "TTTTAAACT"],
            "asv":        ["ASV1", "ASV1"],
        })
        result = get_final_reads(repeated, seqs)
        assert len(result) == 2
        assert list(result["DNAsequence"]) == ["GGGGAAACT", "TTTTAAACT"]

    def test_no_sequences_yields_headers_only(self, long_df):
        result = get_final_reads(long_df, pd.DataFrame(columns=["taxonid"]))
        assert result.empty
        assert "scientificName" in result.columns


class TestAddFinalReads:
    def test_sheet_created(self, monkeypatch, long_df, sequences):
        monkeypatch.setattr(
            "neotoma2faire.write.final_reads.get_taxa_sequences", lambda _: sequences
        )
        wb = openpyxl.Workbook()
        add_final_reads(wb, long_df, 74655)
        ws = wb["finalReads"]
        assert ws.cell(row=1, column=1).value == "scientificName"
        assert ws.cell(row=1, column=3).value == "ASV"
        assert ws.cell(row=2, column=3).value == "ASV1"

    def test_no_sheet_without_sequences(self, monkeypatch, long_df):
        monkeypatch.setattr(
            "neotoma2faire.write.final_reads.get_taxa_sequences",
            lambda _: pd.DataFrame(columns=["taxonid"]),
        )
        wb = openpyxl.Workbook()
        add_final_reads(wb, long_df, 3)
        assert "finalReads" not in wb.sheetnames


class TestFinalReadsColumnOrdering:
    """Sample columns run shallowest-first, depthless last, then natural name."""

    @pytest.fixture
    def scrambled_df(self):
        rows = []
        samples = [
            (1, "WLO50", None, 11),
            (2, "WLO49", 32.5, 22),
            (3, "WLO10", 10.0, 33),
            (4, "WLO9", 9.0, 44),
        ]
        for sampleid, samp_name, depth, reads in samples:
            rows.append({
                "sampleid": sampleid, "samp_name": samp_name,
                "minimumDepthInMeters": depth, "taxonid": 69612,
                "value": reads, "variablename": "Hydrurales", "units": "reads",
            })
        return pd.DataFrame(rows)

    @pytest.fixture
    def one_sequence(self):
        return pd.DataFrame({
            "taxonid": [69612], "taxonname": ["Hydrurales_Clade-II_X_sp."],
            "sequenceid": [12], "sequence": ["GGGGAAACT"], "asv": ["ASV1"],
        })

    def test_sample_columns_are_in_depth_order(self, scrambled_df, one_sequence):
        result = get_final_reads(scrambled_df, one_sequence)
        assert list(result.columns[-4:]) == ["WLO9", "WLO10", "WLO49", "WLO50"]

    def test_counts_stay_with_their_sample(self, scrambled_df, one_sequence):
        result = get_final_reads(scrambled_df, one_sequence)
        assert result.loc[0, "WLO9"] == 44
        assert result.loc[0, "WLO50"] == 11
