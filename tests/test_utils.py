"""Tests for neotoma2faire.utils (formatting, sheet-writing and ordering helpers).

"""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from neotoma2faire.utils import (
    add_sheet_from_dataset,
    checklist_version,
    format_db_value,
    natural_key,
    sort_samples,
    write_flat_sheet,
    write_sheet_rows,
)


class TestFormatDbValue:
    def test_none_returns_empty_string_by_default(self):
        assert format_db_value(None) == ''

    def test_none_returns_custom_placeholder(self):
        assert format_db_value(None, none_placeholder='N/A') == 'N/A'

    def test_scalar_string_returned_unchanged(self):
        assert format_db_value('hello') == 'hello'

    def test_scalar_int_returned_unchanged(self):
        assert format_db_value(42) == 42

    def test_list_joined_with_semicolons(self):
        assert format_db_value(['a', 'b', 'c']) == 'a; b; c'

    def test_list_filters_none_entries(self):
        assert format_db_value(['a', None, 'c']) == 'a; c'

    def test_list_all_none_returns_placeholder(self):
        assert format_db_value([None, None], none_placeholder='empty') == 'empty'

    def test_empty_list_returns_placeholder(self):
        assert format_db_value([], none_placeholder='empty') == 'empty'

    def test_list_of_ints_joined(self):
        assert format_db_value([1, 2, 3]) == '1; 2; 3'


# ---------------------------------------------------------------------------
# Sheet-writing helpers
# ---------------------------------------------------------------------------


def _sheet(columns, header_row=3):
    wb = Workbook()
    ws = wb.active
    ws.title = "target"
    for idx, name in enumerate(columns, start=1):
        ws.cell(row=header_row, column=idx, value=name)
    return wb, ws


def _values(ws, column_name, rows, header_row=3):
    header = {c.value: c.column for c in ws[header_row]}
    return [ws.cell(row=r, column=header[column_name]).value for r in rows]


class TestWriteSheetRows:
    def test_writes_rows_below_the_header(self):
        wb, ws = _sheet(["a", "b"])
        write_sheet_rows(ws, pd.DataFrame({"a": [1, 2], "b": ["x", "y"]}), header_row=3)

        assert _values(ws, "a", (4, 5)) == [1, 2]
        assert _values(ws, "b", (4, 5)) == ["x", "y"]

    def test_sheet_columns_absent_from_the_frame_are_skipped(self):
        wb, ws = _sheet(["a", "unused"])
        write_sheet_rows(ws, pd.DataFrame({"a": [1]}), header_row=3)

        assert _values(ws, "unused", (4,)) == [None]

    def test_frame_columns_absent_from_the_sheet_are_dropped(self):
        wb, ws = _sheet(["a"])
        write_sheet_rows(ws, pd.DataFrame({"a": [1], "extra": [9]}), header_row=3)

        assert _values(ws, "a", (4,)) == [1]

    def test_nan_becomes_none(self):
        wb, ws = _sheet(["a"])
        write_sheet_rows(ws, pd.DataFrame({"a": [float("nan")]}), header_row=3)

        assert _values(ws, "a", (4,)) == [None]

    def test_unwritable_value_falls_back_to_none(self):
        """openpyxl rejects arbitrary objects; the cell is blanked, not crashed."""
        wb, ws = _sheet(["a"])
        write_sheet_rows(ws, pd.DataFrame({"a": [object()]}), header_row=3)

        assert _values(ws, "a", (4,)) == [None]

    def test_list_valued_cell_does_not_crash(self):
        """pd.isna raises on a list; the helper swallows it and writes None."""
        wb, ws = _sheet(["a"])
        write_sheet_rows(ws, pd.DataFrame({"a": [[1, 2]]}), header_row=3)

        assert _values(ws, "a", (4,)) == [None]

    def test_empty_frame_writes_nothing(self):
        wb, ws = _sheet(["a"])
        write_sheet_rows(ws, pd.DataFrame(columns=["a"]), header_row=3)

        assert _values(ws, "a", (4,)) == [None]


class TestWriteFlatSheet:
    def test_creates_a_sheet_with_a_single_header_row(self):
        wb = Workbook()
        write_flat_sheet(wb, "ageModels", pd.DataFrame({"depth": [0.5], "age": [150]}))

        ws = wb["ageModels"]
        assert [c.value for c in ws[1]] == ["depth", "age"]
        assert [c.value for c in ws[2]] == [0.5, 150]

    def test_empty_frame_creates_nothing(self):
        wb = Workbook()
        before = list(wb.sheetnames)

        assert write_flat_sheet(wb, "ageModels", pd.DataFrame()) is None
        assert wb.sheetnames == before

    def test_none_frame_creates_nothing(self):
        wb = Workbook()

        assert write_flat_sheet(wb, "ageModels", None) is None

    def test_existing_sheet_is_reused(self):
        wb = Workbook()
        existing = wb.create_sheet("ageModels")
        ws = write_flat_sheet(wb, "ageModels", pd.DataFrame({"depth": [0.5]}))

        assert ws is existing
        assert wb.sheetnames.count("ageModels") == 1

    def test_after_positions_the_new_sheet(self):
        wb = Workbook()
        wb.active.title = "first"
        wb.create_sheet("second")
        write_flat_sheet(wb, "ageModels", pd.DataFrame({"depth": [0.5]}), after="first")

        assert wb.sheetnames == ["first", "ageModels", "second"]

    def test_unknown_after_appends_at_the_end(self):
        wb = Workbook()
        wb.active.title = "first"
        wb.create_sheet("second")
        write_flat_sheet(wb, "ageModels", pd.DataFrame({"depth": [0.5]}), after="nope")

        assert wb.sheetnames[-1] == "ageModels"


class TestAddSheetFromDataset:
    def test_writes_non_empty_result(self):
        wb, ws = _sheet(["a"])
        df = add_sheet_from_dataset(
            wb, "target", lambda dsid: pd.DataFrame({"a": [dsid]}), 74655
        )

        assert _values(ws, "a", (4,)) == [74655]
        assert list(df["a"]) == [74655]

    def test_empty_result_writes_nothing_and_is_returned(self):
        wb, ws = _sheet(["a"])
        df = add_sheet_from_dataset(wb, "target", lambda dsid: pd.DataFrame(), 74655)

        assert df.empty
        assert _values(ws, "a", (4,)) == [None]

    def test_dataset_id_is_forwarded_to_the_getter(self):
        wb, _ = _sheet(["a"])
        seen = []

        def getter(dsid):
            seen.append(dsid)
            return pd.DataFrame()

        add_sheet_from_dataset(wb, "target", getter, 74655)
        assert seen == [74655]

    def test_custom_header_row(self):
        wb, ws = _sheet(["a"], header_row=1)
        add_sheet_from_dataset(
            wb, "target", lambda dsid: pd.DataFrame({"a": [1]}), 74655, header_row=1
        )

        assert _values(ws, "a", (2,), header_row=1) == [1]


class TestNaturalKey:
    """Sample numbering has to compare numerically, not as text."""

    def test_orders_numbers_numerically(self):
        assert sorted(["WLO10", "WLO9", "WLO1"], key=natural_key) == ["WLO1", "WLO9", "WLO10"]

    def test_is_case_insensitive(self):
        assert natural_key("wlo9") == natural_key("WLO9")

    def test_handles_none_and_bare_text(self):
        assert sorted([None, "b", "a"], key=natural_key) == [None, "a", "b"]


class TestSortSamples:
    """One ordering rule for every sheet: depth, NAs last, then natural name."""

    def _frame(self):
        # Deliberately scrambled, and mirrors West Okoboji: depthless samples
        # come back from the API first, the rest deepest-first.
        return pd.DataFrame(
            {
                "samp_name": ["WLO58", "WLO50", "WLO49", "WLO9", "WLO17"],
                "minimumDepthInMeters": [None, None, 32.5, 8.0, 0.5],
            }
        )

    def test_orders_by_depth_then_puts_missing_depths_last(self):
        out = sort_samples(self._frame())
        assert out["samp_name"].tolist() == ["WLO17", "WLO9", "WLO49", "WLO50", "WLO58"]

    def test_missing_depths_are_ordered_by_natural_name(self):
        out = sort_samples(self._frame())
        assert out["samp_name"].tolist()[-2:] == ["WLO50", "WLO58"]

    def test_index_is_reset(self):
        assert sort_samples(self._frame()).index.tolist() == [0, 1, 2, 3, 4]

    def test_a_missing_depth_column_falls_through_to_name_order(self):
        df = pd.DataFrame({"samp_name": ["WLO10", "WLO9"]})
        assert sort_samples(df)["samp_name"].tolist() == ["WLO9", "WLO10"]

    def test_non_numeric_depths_are_treated_as_missing(self):
        df = pd.DataFrame(
            {"samp_name": ["WLO2", "WLO1"], "minimumDepthInMeters": ["not a depth", 4.0]}
        )
        assert sort_samples(df)["samp_name"].tolist() == ["WLO1", "WLO2"]

    def test_empty_frame_is_returned_unchanged(self):
        df = pd.DataFrame(columns=["samp_name", "minimumDepthInMeters"])
        assert sort_samples(df).empty

    def test_without_the_name_column_it_is_a_no_op(self):
        df = pd.DataFrame({"other": [2, 1]})
        assert sort_samples(df)["other"].tolist() == [2, 1]

    def test_column_names_are_configurable(self):
        df = pd.DataFrame({"Sample": ["b", "a"], "PlotDepth(cm)": [None, 3.0]})
        out = sort_samples(df, name_col="Sample", depth_col="PlotDepth(cm)")
        assert out["Sample"].tolist() == ["a", "b"]


class TestChecklistVersion:
    def test_parses_the_version_from_the_template_name(self):
        assert checklist_version("assets/FAIRe_checklist_v1.0.2.xlsx") == "1.0.2"

    def test_accepts_a_path_object(self):
        assert checklist_version(Path("/tmp/FAIRe_checklist_v2.1.xlsx")) == "2.1"

    def test_returns_none_when_the_name_carries_no_version(self):
        assert checklist_version("template.xlsx") is None
