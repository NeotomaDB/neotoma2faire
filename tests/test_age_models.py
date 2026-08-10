"""Tests for neotoma2faire.extract.age_models and write.age_models."""

import openpyxl
import pandas as pd
import pytest

from neotoma2faire.extract.age_models import get_age_models
from neotoma2faire.write.age_models import add_age_models


@pytest.fixture
def long_df():
    """get_data() frame: two samples, each with a taxon datum and a sed-rate datum."""
    return pd.DataFrame({
        "sampleid":          [1, 1, 2, 2],
        "samp_name":         ["WLO17", "WLO17", "WLO18", "WLO18"],
        "analysisunitname":  ["17-18", "17-18", "18-19", "18-19"],
        "depth":             [0.5, 0.5, 1.5, 1.5],
        "thickness":         [1, 1, 1, 1],
        "age":               [2017.0, 2017.0, 2014.0, 2014.0],
        "agemodel":          ["CRS"] * 4,
        "modelagetype":      ["Calendar years BP"] * 4,
        "variablename":      ["Hydrurales", "Sedimentation rate", "Hydrurales",
                              "Sedimentation rate"],
        "value":             [3860, 0.2629, 6367, 0.3134],
        "units":             ["reads", "cm/yr", "reads", "cm/yr"],
    })


class TestGetAgeModels:
    def test_one_row_per_sample(self, long_df):
        result = get_age_models(long_df)
        assert len(result) == 2
        assert list(result["Sample"]) == ["WLO17", "WLO18"]

    def test_age_column_named_after_chronology(self, long_df):
        result = get_age_models(long_df)
        assert "CRS(Calendar years BP)" in result.columns
        assert list(result["CRS(Calendar years BP)"]) == [2017.0, 2014.0]

    def test_age_column_falls_back_without_chronology(self, long_df):
        df = long_df.drop(columns=["agemodel", "modelagetype", "age"])
        result = get_age_models(df)
        assert "Age" in result.columns
        assert result["Age"].isna().all()

    def test_sed_rate_from_datum(self, long_df):
        result = get_age_models(long_df)
        assert result.loc[0, "SedimentationRate(cm-yr)"] == pytest.approx(0.2629)
        assert result.loc[0, "SedRateUnits"] == "cm/yr"

    def test_depth_and_thickness(self, long_df):
        result = get_age_models(long_df)
        assert list(result["PlotDepth(cm)"]) == [0.5, 1.5]
        assert list(result["Thickness(cm)"]) == [1, 1]
        assert list(result["SampleInterval(cm)"]) == ["17-18", "18-19"]

    def test_chronology_labels_repeated_on_each_row(self, long_df):
        result = get_age_models(long_df)
        assert set(result["AgeModel"]) == {"CRS"}
        assert set(result["AgeType"]) == {"Calendar years BP"}

    def test_empty_frame_yields_headers_only(self):
        result = get_age_models(pd.DataFrame())
        assert result.empty
        assert "Sample" in result.columns


class TestAddAgeModels:
    def test_sheet_created_with_header_on_row_one(self, long_df):
        wb = openpyxl.Workbook()
        add_age_models(wb, long_df)
        ws = wb["ageModels"]
        assert ws.cell(row=1, column=1).value == "Sample"
        assert ws.cell(row=1, column=2).value == "CRS(Calendar years BP)"
        assert ws.cell(row=2, column=1).value == "WLO17"

    def test_no_sheet_when_no_samples(self):
        wb = openpyxl.Workbook()
        add_age_models(wb, pd.DataFrame())
        assert "ageModels" not in wb.sheetnames


class TestAgeModelOrdering:
    """ageModels orders on its own renamed columns, matching the other sheets."""

    @pytest.fixture
    def scrambled_df(self):
        return pd.DataFrame({
            "sampleid":     [1, 2, 3],
            "samp_name":    ["WLO50", "WLO18", "WLO17"],
            "depth":        [None, 1.5, 0.5],
            "thickness":    [1, 1, 1],
            "age":          [None, 2014.0, 2017.0],
            "agemodel":     ["CRS"] * 3,
            "modelagetype": ["Calendar years BP"] * 3,
            "variablename": ["Hydrurales"] * 3,
            "value":        [10, 20, 30],
            "units":        ["reads"] * 3,
        })

    def test_rows_come_out_shallowest_first_with_depthless_last(self, scrambled_df):
        result = get_age_models(scrambled_df)
        assert list(result["Sample"]) == ["WLO17", "WLO18", "WLO50"]

    def test_ages_travel_with_their_samples(self, scrambled_df):
        result = get_age_models(scrambled_df)
        ages = result["CRS(Calendar years BP)"]
        assert list(ages[:2]) == [2017.0, 2014.0]
        assert pd.isna(ages.iloc[2])
