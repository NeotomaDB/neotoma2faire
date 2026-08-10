"""Tests for neotoma2faire.write.experiment_run."""

import pandas as pd
import pytest
from openpyxl import Workbook

from neotoma2faire.write.experiment_run import add_experiment_run


@pytest.fixture
def minimal_workbook():
    """Workbook with an experimentRunMetadata sheet that has a header row at row 3."""
    wb = Workbook()
    ws = wb.active
    ws.title = "experimentRunMetadata"
    # Row 3: column headers matching FAIRe experimentRunMetadata terms
    header_cols = ["samp_name", "assay_name", "pcr_plate_id", "lib_id", "seq_run_id"]
    for col_idx, name in enumerate(header_cols, start=1):
        ws.cell(row=3, column=col_idx, value=name)
    return wb


@pytest.fixture
def sample_df():
    """Minimal long-format DataFrame with two samples."""
    return pd.DataFrame({
        "sampleid":    [1, 1, 2, 2],
        "taxonid":     [10, 20, 10, 20],
        "value":       [5.0, 0.0, 3.0, 1.0],
        "samp_name":   ["Sample_1087", "Sample_1087", "Sample_1088", "Sample_1088"],
        "datasettype": ["pollen surface sample"] * 4,
    })


class TestAddExperimentRun:
    def test_returns_workbook(self, minimal_workbook, sample_df):
        result = add_experiment_run(minimal_workbook, sample_df)
        assert result is minimal_workbook

    def test_writes_one_row_per_sample(self, minimal_workbook, sample_df):
        add_experiment_run(minimal_workbook, sample_df)
        ws = minimal_workbook["experimentRunMetadata"]
        # Row 3 is header; data starts at row 4
        filled_rows = [
            r for r in ws.iter_rows(min_row=4, values_only=True)
            if any(v is not None for v in r)
        ]
        assert len(filled_rows) == 2

    def test_samp_name_written(self, minimal_workbook, sample_df):
        add_experiment_run(minimal_workbook, sample_df)
        ws = minimal_workbook["experimentRunMetadata"]
        # samp_name is column 1
        values = [ws.cell(row=r, column=1).value for r in (4, 5)]
        assert set(values) == {"Sample_1087", "Sample_1088"}

    def test_assay_name_written(self, minimal_workbook, sample_df):
        add_experiment_run(minimal_workbook, sample_df)
        ws = minimal_workbook["experimentRunMetadata"]
        # assay_name is column 2
        value = ws.cell(row=4, column=2).value
        assert value == "pollen surface sample"

    def test_no_samp_name_column_does_not_raise(self, minimal_workbook):
        """DataFrame without samp_name must not crash."""
        df = pd.DataFrame({
            "sampleid": [1],
            "taxonid": [10],
            "value": [5.0],
            "datasettype": ["pollen"],
        })
        add_experiment_run(minimal_workbook, df)  # should not raise

    def test_deduplication(self, minimal_workbook, sample_df):
        """Each sampleid must appear only once even if it has many taxa rows."""
        add_experiment_run(minimal_workbook, sample_df)
        ws = minimal_workbook["experimentRunMetadata"]
        filled_rows = [
            r for r in ws.iter_rows(min_row=4, values_only=True)
            if any(v is not None for v in r)
        ]
        assert len(filled_rows) == 2

    def test_assay_name_from_assay(self, minimal_workbook, sample_df):
        """When an assay is supplied, assay_name comes from assayname, not datasettype."""
        assays = [{"assayname": "18SrRNAV7",
                   "libraries": [{"libid": "L1", "seqrunid": "R1", "pcrplateid": "P1"}]}]
        add_experiment_run(minimal_workbook, sample_df, assays)
        ws = minimal_workbook["experimentRunMetadata"]
        # assay_name col 2; lib fields cols 3-5 (pcr_plate_id, lib_id, seq_run_id)
        assert ws.cell(row=4, column=2).value == "18SrRNAV7"
        assert ws.cell(row=4, column=3).value == "P1"
        assert ws.cell(row=4, column=4).value == "L1"
        assert ws.cell(row=4, column=5).value == "R1"


class TestExperimentRunOrdering:
    """Rows follow the same depth-then-name order as the other sheets."""

    @pytest.fixture
    def scrambled_df(self):
        return pd.DataFrame({
            "sampleid":             [1, 2, 3, 4],
            "samp_name":            ["WLO50", "WLO49", "WLO10", "WLO9"],
            "datasettype":          ["metabarcoding"] * 4,
            "minimumDepthInMeters": [None, 32.5, 10.0, 9.0],
        })

    def test_rows_come_out_shallowest_first_with_depthless_last(
        self, minimal_workbook, scrambled_df
    ):
        add_experiment_run(minimal_workbook, scrambled_df)
        ws = minimal_workbook["experimentRunMetadata"]
        names = [ws.cell(row=r, column=1).value for r in range(4, 8)]
        assert names == ["WLO9", "WLO10", "WLO49", "WLO50"]

    def test_the_ordering_column_is_not_written_to_the_sheet(
        self, minimal_workbook, scrambled_df
    ):
        # minimumDepthInMeters is carried only to sort by; it is not a
        # experimentRunMetadata term and must not leak into the sheet.
        add_experiment_run(minimal_workbook, scrambled_df)
        ws = minimal_workbook["experimentRunMetadata"]
        headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
        assert "minimumDepthInMeters" not in headers
