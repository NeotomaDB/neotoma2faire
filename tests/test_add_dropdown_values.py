"""Tests for neotoma2faire.write.dropdown_values."""

from unittest.mock import MagicMock, patch, call

import pytest
from openpyxl import Workbook

from neotoma2faire.write.dropdown_values import add_dropdown_values, _VOCAB_QUERIES


def _make_mock_conn(vocab_rows: dict[str, list]):
    """Return a mock connection whose cursor returns per-column row lists.

    *vocab_rows* maps SQL query text substrings to the list of 1-dict rows
    that ``fetchall`` should return for that query.
    """
    cur = MagicMock()
    cur.__enter__ = lambda s: s
    cur.__exit__ = MagicMock(return_value=False)

    def fetchall_side_effect():
        # Return the rows registered for the last execute call.
        return cur._last_rows

    def execute_side_effect(query):
        # Match the query against the registered substrings.
        for key, rows in vocab_rows.items():
            if key in query:
                cur._last_rows = rows
                return
        cur._last_rows = []

    cur.execute.side_effect = execute_side_effect
    cur.fetchall.side_effect = fetchall_side_effect

    conn = MagicMock()
    conn.cursor.return_value = cur
    return conn


def _make_workbook(header_cols):
    """Workbook with a Drop-down values sheet whose row 1 has *header_cols*."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Drop-down values"
    for col_idx, name in enumerate(header_cols, start=1):
        ws.cell(row=1, column=col_idx, value=name)
    return wb


class TestAddDropdownValues:
    def test_returns_workbook(self):
        wb = _make_workbook(["assay_type"])
        mock_conn = _make_mock_conn({"assaytypes": [{"assaytype": "targeted"}]})
        with patch("neotoma2faire.write.dropdown_values.neo_connect", return_value=mock_conn):
            result = add_dropdown_values(wb)
        assert result is wb

    def test_writes_vocab_values_under_correct_header(self):
        wb = _make_workbook(["assay_type", "samp_category"])
        vocab = {
            "assaytypes": [{"assaytype": v} for v in ("targeted", "metabarcoding", "other")],
            "samplecategories": [{"samplecategory": "sample"}],
        }
        mock_conn = _make_mock_conn(vocab)
        with patch("neotoma2faire.write.dropdown_values.neo_connect", return_value=mock_conn):
            add_dropdown_values(wb)
        ws = wb["Drop-down values"]
        # assay_type is column 1; values start at row 2
        assert ws.cell(row=2, column=1).value == "targeted"
        assert ws.cell(row=3, column=1).value == "metabarcoding"
        assert ws.cell(row=4, column=1).value == "other"
        # samp_category is column 2
        assert ws.cell(row=2, column=2).value == "sample"

    def test_skips_unknown_header_columns(self):
        """Columns not in _VOCAB_QUERIES must not raise."""
        wb = _make_workbook(["unknown_field"])
        mock_conn = _make_mock_conn({})
        with patch("neotoma2faire.write.dropdown_values.neo_connect", return_value=mock_conn):
            add_dropdown_values(wb)  # should not raise

    def test_handles_empty_vocab_gracefully(self):
        """An empty result from a vocab table must leave the column blank."""
        wb = _make_workbook(["assay_type"])
        mock_conn = _make_mock_conn({"assaytypes": []})
        with patch("neotoma2faire.write.dropdown_values.neo_connect", return_value=mock_conn):
            add_dropdown_values(wb)
        ws = wb["Drop-down values"]
        assert ws.cell(row=2, column=1).value is None

    def test_database_error_leaves_column_unchanged(self):
        """If a query raises (table not yet created) the sheet must not crash."""
        wb = _make_workbook(["assay_type"])
        cur = MagicMock()
        cur.__enter__ = lambda s: s
        cur.__exit__ = MagicMock(return_value=False)
        cur.execute.side_effect = Exception("relation does not exist")
        conn = MagicMock()
        conn.cursor.return_value = cur
        with patch("neotoma2faire.write.dropdown_values.neo_connect", return_value=conn):
            add_dropdown_values(wb)  # should not raise
        ws = wb["Drop-down values"]
        assert ws.cell(row=2, column=1).value is None
