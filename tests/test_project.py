"""Tests for neotoma2faire.write.project."""

import pytest
from openpyxl import Workbook

from neotoma2faire.write import project as project_mod
from neotoma2faire.write.project import add_project


@pytest.fixture(autouse=True)
def _stub_assays(monkeypatch):
    """Default: no assays, so tests don't hit the network. Override per test."""
    monkeypatch.setattr(project_mod, "get_assays_by_dataset", lambda _d: [])


@pytest.fixture
def project_workbook():
    """Workbook with a vertical projectMetadata sheet (term_name in col C)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "projectMetadata"
    ws.cell(1, 1, "requirement_level_code")
    ws.cell(1, 3, "term_name")
    ws.cell(1, 4, "project_level")
    terms = [
        "recordedBy", "project_contact", "project_id",
        "assay_type", "assay_name", "targetTaxonomicAssay", "target_gene",
        "pcr_primer_forward", "pcr_primer_reverse", "platform", "instrument",
        "sterilise_method", "neg_cont_0_1", "pos_cont_0_1", "pcr_0_1", "checkls_ver",
    ]
    for r, term in enumerate(terms, start=2):
        ws.cell(r, 3, term)
    return wb


def _term_value(ws, term):
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 3).value == term:
            return ws.cell(r, 4).value
    return None


class TestAddProject:
    def test_fills_terms_from_endpoint(self, project_workbook, monkeypatch):
        monkeypatch.setattr(project_mod, "get_projects_by_dataset", lambda _d: [{
            "projectname": "WOL_19",
            "participants": [
                {"contactname": "Trisha L. Spanbauer", "email": "trisha@uky.edu"},
                {"contactname": "Jane Doe", "email": "jane@uky.edu"},
            ],
        }])
        add_project(project_workbook, 3)
        ws = project_workbook["projectMetadata"]
        assert _term_value(ws, "project_id") == "WOL_19"
        assert _term_value(ws, "recordedBy") == "Trisha L. Spanbauer; Jane Doe"
        assert _term_value(ws, "project_contact") == "trisha@uky.edu; jane@uky.edu"

    def test_fills_assay_terms(self, project_workbook, monkeypatch):
        monkeypatch.setattr(project_mod, "get_projects_by_dataset", lambda _d: [{
            "projectname": "WOL_19", "participants": [],
        }])
        monkeypatch.setattr(project_mod, "get_dataset", lambda _d: {"datasets": []})
        monkeypatch.setattr(project_mod, "get_assays_by_dataset", lambda _d: [{
            "assayname": "18SrRNAV7",
            "assaytype": "metabarcoding",
            "targettaxonomicassay": "microbial eukaryotes",
            "targetgene": "18S rRNA (SSU eukaryote)",
            "pcrprimerforward": "FWD",
            "pcrprimerreverse": "REV",
            "libraries": [{"platform": "ILLUMINA", "instrument": "Miseq"}],
        }])
        add_project(project_workbook, 74655)
        ws = project_workbook["projectMetadata"]
        assert _term_value(ws, "assay_type") == "metabarcoding"
        assert _term_value(ws, "assay_name") == "18SrRNAV7"
        assert _term_value(ws, "targetTaxonomicAssay") == "microbial eukaryotes"
        assert _term_value(ws, "target_gene") == "18S rRNA (SSU eukaryote)"
        assert _term_value(ws, "pcr_primer_forward") == "FWD"
        assert _term_value(ws, "platform") == "ILLUMINA"
        assert _term_value(ws, "instrument") == "Miseq"

    def test_no_assay_leaves_terms_blank(self, project_workbook, monkeypatch):
        monkeypatch.setattr(project_mod, "get_projects_by_dataset", lambda _d: [{
            "projectname": "WOL_19", "participants": [],
        }])
        monkeypatch.setattr(project_mod, "get_dataset", lambda _d: {"datasets": []})
        add_project(project_workbook, 3)
        ws = project_workbook["projectMetadata"]
        assert _term_value(ws, "project_id") == "WOL_19"
        assert _term_value(ws, "assay_name") is None
        assert _term_value(ws, "target_gene") is None
        assert _term_value(ws, "platform") is None

    def test_project_without_participants_falls_back_to_pi(self, project_workbook, monkeypatch):
        monkeypatch.setattr(project_mod, "get_projects_by_dataset", lambda _d: [{
            "projectname": "WOL_19", "participants": [],
        }])
        monkeypatch.setattr(project_mod, "get_dataset", lambda _d: {
            "datasets": [{"datasetpi": [{"contactname": "Some PI"}]}],
        })
        add_project(project_workbook, 3)
        ws = project_workbook["projectMetadata"]
        assert _term_value(ws, "recordedBy") == "Some PI"
        assert _term_value(ws, "project_contact") is None

    def test_no_project_falls_back_to_pi(self, project_workbook, monkeypatch):
        monkeypatch.setattr(project_mod, "get_projects_by_dataset", lambda _d: [])
        monkeypatch.setattr(project_mod, "get_dataset", lambda _d: {
            "datasets": [{"datasetpi": [{"contactname": "Some PI"}]}],
        })
        add_project(project_workbook, 3)
        ws = project_workbook["projectMetadata"]
        assert _term_value(ws, "recordedBy") == "Some PI"
        assert _term_value(ws, "project_id") is None

    def test_returns_workbook(self, project_workbook, monkeypatch):
        monkeypatch.setattr(project_mod, "get_projects_by_dataset", lambda _d: [])
        monkeypatch.setattr(project_mod, "get_dataset", lambda _d: {"datasets": []})
        assert add_project(project_workbook, 3) is project_workbook


def _assay(**overrides):
    """An assay record shaped like /datasets/{id}/assays returns."""
    assay = {
        "assayname": "18SrRNAV7",
        "assaytype": "metabarcoding",
        "pcrprimerforward": "FWD",
        "pcrprimerreverse": "REV",
        "sterilisemethod": "processed in UToledo clean lab facility",
        "negcont": True,
        "poscont": False,
        "libraries": [],
    }
    assay.update(overrides)
    return assay


@pytest.fixture
def _no_projects(monkeypatch):
    monkeypatch.setattr(project_mod, "get_projects_by_dataset", lambda _d: [])
    monkeypatch.setattr(project_mod, "get_dataset", lambda _d: {"datasets": []})


class TestControlAndSterilisationTerms:
    """The three columns added to ndb.aednaassays reach projectMetadata."""

    def test_sterilise_method_is_written_verbatim(
        self, project_workbook, monkeypatch, _no_projects
    ):
        monkeypatch.setattr(project_mod, "get_assays_by_dataset", lambda _d: [_assay()])
        add_project(project_workbook, 3)
        ws = project_workbook["projectMetadata"]
        assert _term_value(ws, "sterilise_method") == "processed in UToledo clean lab facility"

    def test_booleans_become_the_checklists_1_and_0(
        self, project_workbook, monkeypatch, _no_projects
    ):
        monkeypatch.setattr(project_mod, "get_assays_by_dataset", lambda _d: [_assay()])
        add_project(project_workbook, 3)
        ws = project_workbook["projectMetadata"]
        assert _term_value(ws, "neg_cont_0_1") == "1"
        assert _term_value(ws, "pos_cont_0_1") == "0"

    def test_null_leaves_the_cell_blank_rather_than_asserting_zero(
        self, project_workbook, monkeypatch, _no_projects
    ):
        # The important case: "unknown" must not be written as "no controls used".
        monkeypatch.setattr(
            project_mod, "get_assays_by_dataset",
            lambda _d: [_assay(negcont=None, poscont=None)],
        )
        add_project(project_workbook, 3)
        ws = project_workbook["projectMetadata"]
        assert _term_value(ws, "neg_cont_0_1") is None
        assert _term_value(ws, "pos_cont_0_1") is None


class TestBooleanFlag:
    @pytest.mark.parametrize("value,expected", [
        (True, "1"), (False, "0"),
        (1, "1"), (0, "0"),
        ("true", "1"), ("False", "0"), ("YES", "1"),
        (None, None), ("maybe", None),
    ])
    def test_normalises_every_shape_the_api_might_return(self, value, expected):
        assert project_mod._boolean_flag(value) == expected


class TestPcrDerivation:
    def test_an_amplification_assay_implies_pcr(self, project_workbook, monkeypatch, _no_projects):
        monkeypatch.setattr(project_mod, "get_assays_by_dataset", lambda _d: [_assay()])
        add_project(project_workbook, 3)
        assert _term_value(project_workbook["projectMetadata"], "pcr_0_1") == "1"

    def test_a_recorded_primer_alone_implies_pcr(self):
        assert project_mod._pcr_performed(
            {"assaytype": "something else", "pcrprimerforward": "FWD"}
        ) == "1"

    def test_an_assay_saying_nothing_about_pcr_leaves_it_blank(self):
        assert project_mod._pcr_performed({"assaytype": "shotgun"}) is None

    def test_no_assay_at_all_leaves_it_blank(self, project_workbook, monkeypatch, _no_projects):
        monkeypatch.setattr(project_mod, "get_assays_by_dataset", lambda _d: [])
        add_project(project_workbook, 3)
        assert _term_value(project_workbook["projectMetadata"], "pcr_0_1") is None


class TestChecklistVersionTerm:
    def test_written_when_supplied(self, project_workbook, monkeypatch, _no_projects):
        add_project(project_workbook, 3, checklist_version="1.0.2")
        assert _term_value(project_workbook["projectMetadata"], "checkls_ver") == "1.0.2"

    def test_blank_when_not_supplied(self, project_workbook, monkeypatch, _no_projects):
        add_project(project_workbook, 3)
        assert _term_value(project_workbook["projectMetadata"], "checkls_ver") is None
