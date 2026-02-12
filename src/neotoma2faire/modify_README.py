from openpyxl import load_workbook, Workbook
from datetime import datetime

def modify_README(workbook):
    ws = workbook.active = workbook['README']
    
    ws.insert_rows(3,2)
    ws['A4'] = 'Modified by:'
    ws['A5'] = 'neotoma2FAIRe v0.1.0'
    ws.insert_rows(6,1)
    ws['A8'] = datetime.now()
    return workbook
